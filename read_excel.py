import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('sheet_data_revision.xlsx')

# List all sheet names
print("Sheet names:", workbook.sheetnames)

# Iterate through each sheet and print its content
#for sheet_name in workbook.sheetnames:
#    sheet = workbook[sheet_name]
#    print(f"\nContent of sheet '{sheet_name}':")
#    for row in sheet.iter_rows(values_only=True):
#        print(row)
